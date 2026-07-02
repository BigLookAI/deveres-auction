"""
deVeres Auction — Reconciliation · Export service
===================================================

Serialises reconciliation results to the formats the reviewer needs and to a
clean Odoo-ready intermediate model. Keeping this separate means new output
targets (Odoo importer, PDF) plug in without touching the engine.
"""
from __future__ import annotations

import csv
import io
import json

from .models import DiffStatus, ReconResult, ReconSummary

FLAT_COLUMNS = [
    "buyer_number", "incoming_name", "classification", "recommendation",
    "confidence", "matched_by", "master_ref", "master_name",
    "changed_fields", "action",
]


def to_rows(results: list[ReconResult]) -> list[dict]:
    rows = []
    for r in results:
        rows.append({
            "buyer_number": r.buyer_number,
            "incoming_name": r.incoming_name,
            "classification": r.classification.value,
            "recommendation": r.recommendation.value,
            "confidence": round(r.confidence, 3),
            "matched_by": "|".join(r.matched_by),
            "master_ref": r.master_ref or "",
            "master_name": r.master_name,
            "changed_fields": "|".join(r.changed_fields),
            "action": r.action.value,
        })
    return rows


def to_csv(results: list[ReconResult]) -> str:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=FLAT_COLUMNS)
    w.writeheader()
    for row in to_rows(results):
        w.writerow(row)
    return buf.getvalue()


def to_json(results: list[ReconResult], summary: ReconSummary) -> str:
    return json.dumps({
        "summary": summary.to_dict(),
        "results": [r.to_dict(full=True) for r in results],
    }, indent=2, ensure_ascii=False)


def to_xlsx(results: list[ReconResult], summary: ReconSummary | dict) -> bytes:
    """Excel report: Summary sheet + full Results sheet. Requires openpyxl."""
    from openpyxl import Workbook           # imported lazily — optional dependency
    from openpyxl.styles import Font, PatternFill
    s = summary.to_dict() if hasattr(summary, "to_dict") else dict(summary or {})
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.append(["Metric", "Value"]); ws["A1"].font = ws["B1"].font = Font(bold=True)
    for k in ("total", "new", "retain", "update", "manual_review",
              "ignored_diffs", "avg_confidence", "master_records", "processing_ms"):
        ws.append([k.replace("_", " ").title(), s.get(k, "")])
    ws2 = wb.create_sheet("Results"); ws2.append([c.replace("_", " ").title() for c in FLAT_COLUMNS])
    for c in ws2[1]:
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="EEF3F0")
    for row in to_rows(results):
        ws2.append([row[c] for c in FLAT_COLUMNS])
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def to_pdf_summary(results: list[ReconResult], summary: ReconSummary | dict) -> bytes:
    """One-page PDF summary: counts, confidence, decisions, top changes. Requires fpdf2."""
    from fpdf import FPDF                   # imported lazily — optional dependency
    s = summary.to_dict() if hasattr(summary, "to_dict") else dict(summary or {})
    by_action: dict[str, int] = {}
    for r in results:
        by_action[r.action.value] = by_action.get(r.action.value, 0) + 1
    pdf = FPDF(); pdf.set_auto_page_break(True, 18); pdf.add_page()
    pdf.set_font("Helvetica", "B", 16); pdf.set_text_color(11, 107, 80)
    pdf.cell(0, 10, "deVeres Auction - Contact Reconciliation Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9); pdf.set_text_color(93, 107, 100)
    pdf.cell(0, 6, "Blue Cubes upload reconciled against the canonical client database - by Cimelium",
             new_x="LMARGIN", new_y="NEXT"); pdf.ln(3)
    pdf.set_text_color(24, 33, 29)
    def row(label, value, bold=False):
        pdf.set_font("Helvetica", "B" if bold else "", 11)
        pdf.cell(90, 8, label); pdf.cell(0, 8, str(value), new_x="LMARGIN", new_y="NEXT")
    row("Total uploaded contacts", s.get("total", 0), bold=True)
    row("New (ADD)", s.get("new", 0)); row("Existing (KEEP EXISTING)", s.get("retain", 0))
    row("Updates suggested", s.get("update", 0)); row("Manual review", s.get("manual_review", 0))
    row("Formatting-only differences ignored", s.get("ignored_diffs", 0))
    row("Average match confidence", f"{float(s.get('avg_confidence', 0))*100:.1f}%")
    row("Master records compared against", f"{s.get('master_records', 0):,}")
    row("Processing time", f"{s.get('processing_ms', 0)} ms"); pdf.ln(3)
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "Reviewer decisions", new_x="LMARGIN", new_y="NEXT")
    for k, v in sorted(by_action.items()):
        row(f"  {k}", v)
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 12); pdf.cell(0, 8, "Update-suggested contacts (top 20)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    for r in [x for x in results if x.classification.value == "update"][:20]:
        pdf.cell(0, 6, f"  {r.incoming_name}  (#{r.buyer_number})  ->  {', '.join(r.changed_fields) or '-'}",
                 new_x="LMARGIN", new_y="NEXT")
    out = pdf.output()
    return bytes(out)


def odoo_intermediate(results: list[ReconResult]) -> list[dict]:
    """Clean intermediate model, ready to feed a future Odoo importer.

    Each entry states the action, the canonical record (source of truth), the
    incoming record, and the difference report — so the importer needs no
    reconciliation logic of its own.
    """
    out = []
    for r in results:
        out.append({
            "action": r.action.value,                 # ADD | UPDATE | IGNORE | MANUAL_REVIEW
            "confidence": round(r.confidence, 3),
            "matched_by": r.matched_by,
            "canonical_record": r.master or None,      # None for NEW
            "canonical_ref": r.master_ref,
            "incoming_record": r.incoming,
            "buyer_number": r.buyer_number,
            "difference_report": [
                {"field": d.field, "current": d.current, "incoming": d.incoming,
                 "status": d.status.value, "significant": d.significant}
                for d in r.diffs if d.status != DiffStatus.UNCHANGED
            ],
            "lots": r.lots,
        })
    return out
